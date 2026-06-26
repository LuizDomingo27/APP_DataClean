"""
Service: exportação dos DataFrames para arquivos .xlsx com formatação profissional.
"""

from __future__ import annotations

import io
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path


# ─── Paleta de cores ──────────────────────────────────────────────────────────
HEADER_BG   = "1E2A3A"   # azul-marinho escuro
HEADER_FG   = "00E5FF"   # ciano elétrico
ROW_ALT     = "F0F4F8"   # cinza-azulado claro para linhas alternadas
BORDER_CLR  = "CBD5E1"

_THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_CLR),
    right=Side(style="thin", color=BORDER_CLR),
    top=Side(style="thin", color=BORDER_CLR),
    bottom=Side(style="thin", color=BORDER_CLR),
)


def _format_sheet(ws):
    """Aplica formatação visual ao worksheet."""
    # Cabeçalho
    for cell in ws[1]:
        cell.font      = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER

    # Linhas de dados
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bg = ROW_ALT if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border    = _THIN_BORDER
            cell.font      = Font(name="Calibri", size=10)

            # Formatos específicos por tipo
            if isinstance(cell.value, bool):
                pass
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, (datetime.datetime, datetime.date)):
                cell.number_format = "DD/MM/YYYY"

    # Largura automática
    def _display_len(v) -> int:
        if v is None:
            return 0
        if isinstance(v, (datetime.datetime, datetime.date)):
            return 10  # comprimento de "DD/MM/YYYY"
        return len(str(v))

    for col_cells in ws.columns:
        max_len = max(
            (_display_len(c.value) for c in col_cells),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 4, 40
        )

    # Altura do cabeçalho
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Dados") -> bytes:
    """Serializa um DataFrame em bytes xlsx com formatação."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="DD/MM/YYYY", date_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _format_sheet(ws)
    return buf.getvalue()


def save_all(results: dict[str, pd.DataFrame], output_dir: Path) -> dict[str, Path]:
    """
    Salva todos os DataFrames em disco.

    `results` é um dict {nome_arquivo: DataFrame}.
    Retorna dict {nome_arquivo: Path_salvo}.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    saved: dict[str, Path] = {}

    for name, df in results.items():
        if df is None or df.empty:
            continue
        path = output_dir / f"{name.upper()}.xlsx"
        data = df_to_excel_bytes(df, sheet_name=name[:31])
        path.write_bytes(data)
        saved[name] = path

    return saved
